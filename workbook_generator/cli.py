"""Write only the three allowed forecast cells in a supplied workbook template."""
from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path
from typing import Any

import openpyxl


class WorkbookGenerationError(ValueError):
    pass


def _normalise(value: Any) -> str:
    return "" if value is None else str(value).strip().casefold()


def _header_row(sheet, target_period: str) -> int:
    for row in range(1, 31):
        if (
            _normalise(sheet.cell(row, 1).value) == "metric"
            and _normalise(sheet.cell(row, 2).value) == "units"
            and _normalise(sheet.cell(row, 3).value) == _normalise(target_period)
        ):
            return row
    raise WorkbookGenerationError(f"Summary header Metric/Units/{target_period} not found")


def write_workbook(
    *, company_id: str, forecast: dict[str, Any], template: str | Path, output: str | Path,
) -> Path:
    if forecast.get("schema_version") != "company_forecast.v1":
        raise WorkbookGenerationError("forecast must use company_forecast.v1")
    if forecast.get("company_id") != company_id:
        raise WorkbookGenerationError("forecast company_id does not match requested company")
    target_period = forecast.get("target_period")
    if not isinstance(target_period, str) or not target_period:
        raise WorkbookGenerationError("forecast target_period is required")
    values: dict[tuple[str, str], float] = {}
    for item in forecast.get("forecasts", []):
        if not isinstance(item, dict):
            raise WorkbookGenerationError("forecast entries must be objects")
        label, units, value = item.get("metric"), item.get("units"), item.get("value")
        if not isinstance(label, str) or not isinstance(units, str):
            raise WorkbookGenerationError("forecast metric and units are required")
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise WorkbookGenerationError(f"forecast value for {label} is not numeric")
        key = (_normalise(label), _normalise(units))
        if key in values:
            raise WorkbookGenerationError(f"duplicate forecast for {label} / {units}")
        values[key] = float(value)
    if len(values) != 3:
        raise WorkbookGenerationError("forecast must contain exactly three unique metrics")

    template_path, output_path = Path(template), Path(output)
    if not template_path.is_file():
        raise WorkbookGenerationError(f"template does not exist: {template_path}")
    workbook = openpyxl.load_workbook(template_path)
    if "Summary" not in workbook.sheetnames:
        raise WorkbookGenerationError("template has no Summary sheet")
    sheet = workbook["Summary"]
    header = _header_row(sheet, target_period)
    written: set[tuple[str, str]] = set()
    for row in range(header + 1, header + 4):
        key = (_normalise(sheet.cell(row, 1).value), _normalise(sheet.cell(row, 2).value))
        if key not in values:
            raise WorkbookGenerationError(f"template metric has no matching forecast: {sheet.cell(row, 1).value}")
        sheet.cell(row, 3).value = values[key]
        written.add(key)
    if written != set(values):
        raise WorkbookGenerationError("forecast contains a metric absent from the template")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)
    workbook.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Write a validated company forecast into one workbook template")
    parser.add_argument("--company", required=True)
    parser.add_argument("--forecast", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    payload = json.loads(Path(args.forecast).read_text(encoding="utf-8"))
    destination = write_workbook(
        company_id=args.company, forecast=payload, template=args.template, output=args.output,
    )
    print(json.dumps({"company": args.company, "output": str(destination)}))


if __name__ == "__main__":
    main()
